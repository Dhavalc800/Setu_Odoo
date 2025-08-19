from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import csv
import io
import xlrd
from openpyxl import load_workbook
from odoo.tools.translate import _
import logging
from openpyxl import Workbook
from odoo.tools import email_normalize

class CampaignsUploadWizard(models.TransientModel):
    _name = 'campaigns.upload.wizard'
    _description = 'Upload Users via CSV or Excel'

    file = fields.Binary(string='File')
    file_name = fields.Char(string='Filename')
    campaign_id = fields.Many2one('campaigns.list', string='Campaign', required=True)
    select_method_upd = fields.Selection(
        related='campaign_id.select_method_upd',
        store=False,
    )

    def export_user_email_csv(self):
        """Generate CSV file with user data"""
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_ALL)
        
        # Write header
        writer.writerow(['User Name', 'User Email'])
        
        # Prepare file for download
        output.seek(0)
        file_data = output.getvalue().encode('utf-8')
        output.close()
        
        return self._prepare_download(file_data, 'csv')

    def export_user_email_excel(self):
        """Generate Excel file with user data"""
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "User Data"
        
        # Write header
        worksheet.append(['User Name', 'User Email'])
        
        # Save workbook to BytesIO
        workbook.save(output)
        output.seek(0)
        file_data = output.getvalue()
        output.close()
        
        return self._prepare_download(file_data, 'xlsx')

    def import_file(self):
        if not self.file_name:
            raise UserError(_("Please upload a valid CSV or Excel file."))
        
        if self.file_name.endswith('.csv'):
            self._import_csv()
        elif self.file_name.endswith('.xls') or self.file_name.endswith('.xlsx'):
            self._import_excel()
        else:
            raise UserError(_("Unsupported file format."))

    def _import_csv(self):
        # Decode base64 and parse CSV
        data = base64.b64decode(self.file)
        csv_reader = csv.reader(io.StringIO(data.decode("utf-8-sig")))  # Handles UTF-8 with BOM

        # Read and validate header
        raw_header = next(csv_reader)
        header = [h.strip() for h in raw_header if h.strip()]
        expected_header = ['User Name', 'User Email']
        if header != expected_header:
            raise UserError(_(
                "Invalid CSV header. Expected: 'User Name, User Email'. Got: %s" % raw_header
            ))

        # Collect valid entries
        email_to_name = {}
        for row in csv_reader:
            cleaned = [r.strip() for r in row]
            if len(cleaned) < 2 or not cleaned[0] or not cleaned[1]:
                continue
            name, email = cleaned[0], cleaned[1]
            email_to_name[email] = name

        if not email_to_name:
            raise UserError(_("No valid user entries found in the file."))

        # Lookup users by email
        user_model = self.env['res.users']
        existing_users = user_model.search([('email', 'in', list(email_to_name.keys()))])
        user_ids = existing_users.ids

        if not user_ids:
            raise UserError(_("No matching users found in the system for given emails."))

        # Optional: Show emails not found
        not_found_emails = set(email_to_name.keys()) - set(existing_users.mapped('email'))

        # Combine with existing users already linked to campaign
        existing_user_ids = self.campaign_id.user_ids.ids
        combined_user_ids = list(set(existing_user_ids + user_ids))

        # Update campaign user_ids without removing old users
        self.campaign_id.user_ids = [(6, 0, combined_user_ids)]

        # Optional: success message
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Import Complete"),
                'message': _("Users successfully imported and assigned to the campaign."),
                'type': 'success',
                'sticky': False,
            }
        }

    # def _import_csv(self):
    #     data = base64.b64decode(self.file)
    #     csv_reader = csv.reader(io.StringIO(data.decode("utf-8-sig")))  # Handle UTF-8 with BOM

    #     # Read and clean header
    #     raw_header = next(csv_reader)
    #     header = [h.strip() for h in raw_header if h.strip()]
    #     expected_header = ['User Name', 'User Email']

    #     if header != expected_header:
    #         raise UserError(_(
    #             "Invalid CSV header. Expected: 'User Name,User Email'. Got: %s" % raw_header
    #         ))

    #     # Use a set to avoid duplicates and collect all emails
    #     email_to_name = {}
    #     for row in csv_reader:
    #         cleaned = [r.strip() for r in row]
    #         if len(cleaned) < 2 or not cleaned[0] or not cleaned[1]:
    #             continue
    #         name, email = cleaned[0], cleaned[1]
    #         email_to_name[email] = name

    #     if not email_to_name:
    #         raise UserError(_("No valid user entries found in the file."))

    #     # Bulk search users by email
    #     user_model = self.env['res.users']
    #     existing_users = user_model.search([('email', 'in', list(email_to_name.keys()))])
    #     user_ids = existing_users.ids

    #     if not user_ids:
    #         raise UserError(_("No matching users found in the system for given emails."))

    #     # Link users to campaign
    #     self.campaign_id.user_ids = [(6, 0, user_ids)]

    def _import_excel(self):
        # Decode and load the Excel file
        data = base64.b64decode(self.file)
        workbook = load_workbook(io.BytesIO(data), read_only=True)
        sheet = workbook.active

        # Read and validate header
        raw_header = [cell.value.strip() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header = [h for h in raw_header if h]
        expected_header = ['User Name', 'User Email']
        if header != expected_header:
            raise UserError(_(
                "Invalid Excel header. Expected: 'User Name, User Email'. Got: %s" % raw_header
            ))

        # Collect name and email data
        email_to_name = {}
        for row in sheet.iter_rows(min_row=2):
            name = row[0].value.strip() if row[0].value else ''
            email = row[1].value.strip() if row[1].value else ''
            if name and email:
                email_to_name[email] = name

        if not email_to_name:
            raise UserError(_("No valid user entries found in the file."))

        # Lookup users by email
        user_model = self.env['res.users']
        existing_users = user_model.search([('email', 'in', list(email_to_name.keys()))])
        user_ids = existing_users.ids

        if not user_ids:
            raise UserError(_("No matching users found in the system for given emails."))

        # Optional: Warn if some emails not found
        not_found_emails = set(email_to_name.keys()) - set(existing_users.mapped('email'))
        # if not_found_emails:
        #     raise UserError(_("Some users not found in system: %s" % ', '.join(not_found_emails)))

        # Combine with existing assigned users
        existing_user_ids = self.campaign_id.user_ids.ids
        combined_user_ids = list(set(existing_user_ids + user_ids))

        # Assign updated list to campaign
        self.campaign_id.user_ids = [(6, 0, combined_user_ids)]

        # Optional: Success Message (if called from button)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Success"),
                'message': _("Users successfully imported and assigned to the campaign."),
                'type': 'success',
                'sticky': False,
            }
        }


    # def _import_excel(self):
    #     data = base64.b64decode(self.file)
    #     workbook = load_workbook(io.BytesIO(data), read_only=True)
    #     sheet = workbook.active

    #     # Read header
    #     raw_header = [cell.value.strip() if cell.value else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    #     header = [h for h in raw_header if h]
    #     expected_header = ['User Name', 'User Email']

    #     if header != expected_header:
    #         raise UserError(_(
    #             "Invalid Excel header. Expected: 'User Name,User Email'. Got: %s" % raw_header
    #         ))

    #     # Collect all emails and names
    #     email_to_name = {}
    #     for row in sheet.iter_rows(min_row=2):
    #         name = row[0].value.strip() if row[0].value else ''
    #         email = row[1].value.strip() if row[1].value else ''
    #         if name and email:
    #             email_to_name[email] = name

    #     if not email_to_name:
    #         raise UserError(_("No valid user entries found in the file."))

    #     # Bulk user lookup
    #     user_model = self.env['res.users']
    #     existing_users = user_model.search([('email', 'in', list(email_to_name.keys()))])
    #     user_ids = existing_users.ids
    #     print("\n\n\nUser IDs:", user_ids)
    #     if not user_ids:
    #         raise UserError(_("No matching users found in the system for given emails."))

    #     # Assign users to campaign
    #     self.campaign_id.user_ids = [(6, 0, user_ids)]
    
    def _prepare_download(self, file_data, file_ext):
        """Create download action for the generated file"""
        file_name = f"campaign_users_{self.campaign_id.name}.{file_ext}"
        
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'datas': base64.b64encode(file_data),
            'res_model': self._name,
            'type': 'binary',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }